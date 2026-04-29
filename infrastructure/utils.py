from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



#PDF

def render_to_pdf(template_src, context_dict=None):
    if context_dict is None:
        context_dict = {}

    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()

    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        return result.getvalue()
    return None



#ESTILO EXCEL

def get_excel_styles():
    return {
        "header_font": Font(bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid"),
        "center": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }



#HISTORIAL EXCEL

def build_history_excel(tickets):
    styles = get_excel_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    #Historial
    headers = ["Placa", "Tipo", "Propietario", "Espacio", "Entrada", "Salida", "Total", "Estado"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    row = 2

    for t in tickets:
        estado = "En Curso" if t.status == "ACTIVE" else "Finalizado"

        ws.append([
            t.vehicle.license_plate,
            t.vehicle.get_type_display() if hasattr(t.vehicle, "get_type_display") else t.vehicle.type,
            t.vehicle.client.name if t.vehicle.client else "Visitante",
            f"#{t.parking_spot.number}" if t.parking_spot else "--",
            t.entry_time.strftime("%d/%m/%Y %H:%M") if t.entry_time else "",
            t.exit_time.strftime("%d/%m/%Y %H:%M") if t.exit_time else "--:--",
            float(t.total_paid) if t.total_paid else 0,
            estado
        ])

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]

            if col in [2, 4, 8]:
                cell.alignment = styles["center"]

        #Ingreso
        ws.cell(row=row, column=7).number_format = '"$"#,##0'

        #Estado
        estado_cell = ws.cell(row=row, column=8)

        if t.status == "ACTIVE":
            estado_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            estado_cell.font = Font(bold=True, color="166534")
        else:
            estado_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            estado_cell.font = Font(bold=True, color="475569")

        row += 1

    widths = [15, 12, 25, 10, 20, 20, 15, 15]

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    return wb